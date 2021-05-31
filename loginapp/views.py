from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login,logout
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.forms.models import model_to_dict
import datetime
import os
import re
from django import forms
from .forms import EmployeeRecordForm,EmployeePersonalForm,EmergencyForm,JobsForm, BranchForm, SpouseForm, EmploymentHistoryForm,EducationalBackgroundForm,FamilyMemberBackgroundForm,ChildBackgroundForm, EmployeeDocument, EditRecordForm, AwardsRecord, RecordDocument, AwardsRecordEdit,EditDocument,DisciplineRecord,DisciplineRecordEdit
from .models import Employee, EmployeePersonalInfo,EmployeePosition,EmployeeWorkLocation, ChildBackground,SpouseBackground,FamilyMemberBackground,EducationalBackground,EmploymentHistory, EmergencyDetails, Document, Record, Awards, Noac,Awol,Timekeeping ,MajorOffense
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, Http404
from django.forms import formset_factory
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.writer.excel import save_virtual_workbook
from django.utils.encoding import smart_str
from formtools.wizard.views import SessionWizardView
from itertools import chain
from operator import itemgetter

    # Redirect to a success page.
def checkmodel(classmodel, **kwargs):
    try:
        return classmodel.objects.get(**kwargs)
    except classmodel.DoesNotExist:
        return None

def checkmodelq(classmodel, arg):
    try:
        return classmodel.objects.get(arg)
    except classmodel.DoesNotExist:
        return None

# Create your views here.
@login_required
def employeeform(request):
    employees = Employee.objects.filter(deletehide = 0).order_by('employeeid')
    documents = Document.objects.filter(employeeid__deletehide = 0).filter(documenthide = 0)
    employeedocu = Document.objects.filter(Q(memoreferencenumber = None) , Q(employeeid__deletehide=0) , Q(documenthide=0))
    awards = Document.objects.filter(Q(memoreferencenumber__recordtype='Award'),Q(employeeid__deletehide=0), Q(documenthide=0))
    discipline = Document.objects.filter(Q(memoreferencenumber__recordtype='Discipline'), Q(employeeid__deletehide=0), Q(documenthide=0))
    context = {
    'employees': employees,
    'count': employees.count(),
    'documentcount' : documents.count(),
    'employeedocucount' : employeedocu.count(),
    'awardscount' : awards.count(),
    'disciplinecount' : discipline.count(),
    'documents': documents
    }
    if request.method == "POST":
        sortbytype = request.POST.get('sorted')
        searchempid = request.POST.get('searchempid')
        searchempname = request.POST.get('searchempname')
        searchbranch = request.POST.get('searchbranch')
        searchregion = request.POST.get('searchregion')
        searchdept = request.POST.get('searchdept')
        searchposition = request.POST.get('searchposition')
        searchempstatus = request.POST.get('searchempstatus')
        employees = Employee.objects.filter(deletehide = 0)

        if searchempid != '' or searchempid != None:
            if searchempid.isdecimal():
                employees = employees.filter(employeeid = int(searchempid))


        if searchempname != '' :
            if searchempname.replace(' ', '').replace('.',' ').isalpha(): ## This other condition can be removed, as long as it isn't empty is what it matters
                employees = employees.filter(informationid__employeename__icontains = searchempname)


        if searchbranch != '' :
            if searchbranch.replace(' ', '').replace('.',' ').isalpha():
                employees = employees.filter(branch__branch__icontains = searchbranch)


        if searchregion != '' :
            if searchregion.replace(' ', '').replace('.',' ').isalpha():
                employees = employees.filter(branch__region__icontains = searchregion)


        if searchdept != '':
            if searchdept.replace(' ', '').replace('.',' ').isalpha():
                employees = employees.filter(jobid__department__icontains = searchdept)


        if searchposition != '':
            if searchposition.replace(' ', '').replace('.',' ').isalpha():
                employees = employees.filter(jobid__position__icontains = searchposition)

        if searchempstatus != '':
                employees = employees.filter(employmentstatus__icontains = searchempstatus)

        if sortbytype == 'employeename':
            employees = employees.order_by('informationid__employeename')
        elif sortbytype == 'branch' or sortbytype == 'region':
            employees = employees.order_by('branch__' + sortbytype)
        elif sortbytype == 'department' or sortbytype == 'position':
            employees = employees.order_by('jobid__' + sortbytype)
        else:
            employees = employees.order_by(sortbytype)

        context['employees'] = employees
        context['count'] = employees.count()

        return render(request, 'loginapp/employee_masterlist.html',context)
    return render(request, 'loginapp/employee_masterlist.html',context)

@login_required
def employeeprof(request,empid):
    employee = get_object_or_404(Employee, employeeid=empid)
    information = employee.informationid
    emergency = information.emergencycontactnumber
    spouse = information.spouseid
    emphistory = EmploymentHistory.objects.filter(informationid=information)
    education = EducationalBackground.objects.filter(informationid=information)
    family = FamilyMemberBackground.objects.filter(informationid=information)
    children = ChildBackground.objects.filter(informationid=information)
    employeedocu = Document.objects.filter(Q(memoreferencenumber = None) , Q(employeeid__deletehide=0) , Q(documenthide=0))
    degree = None
    if education.count() >= 1:
        degree = education[0].highestdegree
    context = {
        'employee': employee,
        'information': information,
        'emergency': emergency,
        'spouse': spouse,
        'degree': degree,
        'emphistory' :emphistory,
        'education':education,
        'family' :family,
        'children':children,
        'documents': employeedocu,

    }
    return render(request, 'loginapp/prof.html', context)

@login_required
def employeedelete(request,empid):
    employee = get_object_or_404(Employee, employeeid=empid)
    employee.deletehide = 1
    employee.save()





    return redirect('employeeform') #change for final product





@login_required
def documentsdelete(request,did):
    document = get_object_or_404(Document, documentid=did)

    document.documenthide = 1
    document.save()
    return redirect('viewtest_home')


@login_required
def editrecord(request,empid):
    employee = get_object_or_404(Employee, employeeid=empid)
    requiredData = {
    'employeeid': employee.employeeid,
    'startdate': employee.startdate,
    'enddate' : employee.enddate,
    'employmentstatus' : employee.employmentstatus,
    'salarytype' : employee.salarytype,
    'salary': employee.salary,
    'branch': employee.branch,
    'jobid': employee.jobid,
    }
    recordform = EditRecordForm(requiredData)
    if request.method == "POST":
        recordform = EditRecordForm(request.POST)
        if recordform.is_valid():
            employee.startdate = recordform.cleaned_data['startdate']
            employee.enddate = recordform.cleaned_data['enddate']
            employee.employmentstatus = recordform.cleaned_data['employmentstatus']
            employee.salarytype = recordform.cleaned_data['salarytype']
            employee.salary = recordform.cleaned_data['salary']
            employee.branch = recordform.cleaned_data['branch']
            employee.jobid = recordform.cleaned_data['jobid']
            employee.save()
            return redirect('/home/employee/' + str(empid))
    recordform.fields['employeeid'].widget.attrs['readonly'] = True
    context = {
    'form': recordform,}
    return render(request, 'loginapp/edit/1.html',context)



@login_required
def editpersonal(request,empid):
    employee = get_object_or_404(Employee, employeeid=empid)
    information = employee.informationid
    requiredData = {
    'employeename'  : information.employeename,
    'gender' : information.gender,
    'birthdate' : information.birthdate,
    'civilstatus' : information.civilstatus,
    'citizenship' : information.citizenship,
    'religion' : information.religion,
    'bloodtype' : information.bloodtype,
    'numberofdependent' : information.numberofdependent,
    'presentaddress' : information.presentaddress,
    'permanentaddress' : information.permanentaddress,
    'contactnumber' : information.contactnumber,
    }
    recordform = EmployeePersonalForm(requiredData)
    if request.method == "POST":
        recordform = EmployeePersonalForm(request.POST)
        if recordform.is_valid():
            information.employeename= recordform.cleaned_data['employeename']
            information.gender = recordform.cleaned_data['gender']
            information.birthdate = recordform.cleaned_data['birthdate']
            information.civilstatus = recordform.cleaned_data['civilstatus']
            information.citizenship = recordform.cleaned_data['citizenship']
            information.religion = recordform.cleaned_data['religion']
            information.bloodtype = recordform.cleaned_data['bloodtype']
            information.numberofdependent = recordform.cleaned_data['numberofdependent']
            information.presentaddress = recordform.cleaned_data['presentaddress']
            information.permanentaddress = recordform.cleaned_data['permanentaddress']
            information.contactnumber = recordform.cleaned_data['contactnumber']
            information.save()
            return redirect('/home/employee/' + str(empid))
    context = {
    'form': recordform,}
    return render(request, 'loginapp/edit/2.html',context)

@login_required
def editemergency(request,empid):
    employee = get_object_or_404(Employee, employeeid=empid)
    information = employee.informationid
    emergency = EmergencyDetails.objects.get(pk = information.emergencycontactnumber.emergencycontactnumber)
    requiredData = {
    'emergencycontactnumber' : emergency.emergencycontactnumber,
    'emergencycontactname' : emergency.emergencycontactname,
    'emergencyrelationship' : emergency.emergencyrelationship,
    'emergencyaddress' : emergency.emergencyaddress,
    }
    recordform = EmergencyForm(requiredData)
    if request.method == "POST":
        recordform = EmergencyForm(request.POST)
        if recordform.is_valid():
            if checkmodel(EmergencyDetails, emergencycontactnumber = recordform.cleaned_data['emergencycontactnumber']) != emergency and checkmodel(EmergencyDetails, emergencycontactnumber = recordform.cleaned_data['emergencycontactnumber']) != None:
                new = EmergencyDetails.objects.get(emergencycontactnumber = recordform.cleaned_data['emergencycontactnumber'])
                new.emergencycontactname= recordform.cleaned_data['emergencycontactname']
                new.emergencyrelationship= recordform.cleaned_data['emergencyrelationship']
                new.emergencyaddress=recordform.cleaned_data['emergencyaddress']
                information.emergencycontactnumber = new
                information.save()

                if not EmployeePersonalInfo.objects.filter(emergencycontactnumber = emergency).exists():
                    emergency.delete()
            elif recordform.cleaned_data['emergencycontactnumber'] != emergency.emergencycontactnumber:
                new = EmergencyDetails.objects.create(
                    emergencycontactnumber = recordform.cleaned_data['emergencycontactnumber'],
                    emergencycontactname= recordform.cleaned_data['emergencycontactname'],
                    emergencyrelationship= recordform.cleaned_data['emergencyrelationship'],
                    emergencyaddress=recordform.cleaned_data['emergencyaddress'])
                information.emergencycontactnumber = new
                information.save()

                if not EmployeePersonalInfo.objects.filter(emergencycontactnumber = emergency).exists():
                    emergency.delete()
            else:
                emergency.emergencycontactname= recordform.cleaned_data['emergencycontactname']
                emergency.emergencyrelationship= recordform.cleaned_data['emergencyrelationship']
                emergency.emergencyaddress=recordform.cleaned_data['emergencyaddress']
                emergency.save()
                information.save()
            return redirect('/home/employee/' + str(empid))
    context = {
    'form': recordform,}
    return render(request, 'loginapp/edit/3.html',context)


@login_required
def editemphistory(request,empid):
    employee = get_object_or_404(Employee, employeeid=empid)
    information = employee.informationid
    emphistory = EmploymentHistory.objects.filter(informationid=information)
    emphistoryData = []
    counts = 1
    for history in emphistory:
        counts = 0
        dic = {
            'previouscompanyname': history.previouscompanyname,
            'previousposition'  : history.previousposition,
            'reasonforleaving'   : history.reasonforleaving,
            'companycontactnumber' : history.companycontactnumber,
            'withcoeorclearance' : history.withcoeorclearance,
        }
        emphistoryData.append(dic)
    emphistoryform1 = formset_factory(EmploymentHistoryForm, extra = counts )
    if request.method == "POST":
        emphistoryform = emphistoryform1(request.POST,initial = emphistoryData)
        if emphistoryform.is_valid():
            counts = [0,0,0,0]
            totals = [0,0,0,0]
            for histories in emphistoryform:
                totals[0] += 1

            if totals[0] < emphistory.count():
                #print(totals[0] )
                for x in range(emphistory.count() - 1, totals[0] - 1, -1):

                    todel = emphistory[x]
                    todel.delete()

            for histories in emphistoryform:
                counts[0] += 1
                if counts[0] <= emphistory.count():
                    curhistory = emphistory[counts[0] - 1]
                    if histories['previouscompanyname'].value() and histories['withcoeorclearance'].value():
                        curhistory.previouscompanyname= histories.cleaned_data['previouscompanyname']
                        curhistory.previousposition= histories.cleaned_data['previousposition']
                        curhistory.reasonforleaving= histories.cleaned_data['reasonforleaving']
                        curhistory.companycontactnumber= histories.cleaned_data['companycontactnumber']
                        curhistory.withcoeorclearance= histories.cleaned_data['withcoeorclearance']
                        curhistory.save()
                else:
                    if histories['previouscompanyname'].value() and histories['withcoeorclearance'].value():
                        EmploymentHistory.objects.create(
                        informationid=information,
                        previouscompanyname= histories.cleaned_data['previouscompanyname'],
                        previousposition= histories.cleaned_data['previousposition'],
                        reasonforleaving= histories.cleaned_data['reasonforleaving'],
                        companycontactnumber= histories.cleaned_data['companycontactnumber'],
                        withcoeorclearance= histories.cleaned_data['withcoeorclearance']
                        )
            return redirect('/home/employee/' + str(empid))
    else:
        emphistoryform = emphistoryform1(initial = emphistoryData)
    context = {
    'form': emphistoryform,}
    return render(request, 'loginapp/edit/4.html',context)

@login_required
def editeducation(request,empid):
    employee = get_object_or_404(Employee, employeeid=empid)
    information = employee.informationid
    education = EducationalBackground.objects.filter(informationid=information)
    educationData = []
    counts = 1
    for background in education:
        counts = 0
        dic = {
            'highestdegree': background.highestdegree,
            'schoolname'  : background.schoolname,
            'startingyearattended'   : background.startingyearattended,
            'endingyearattended' : background.endingyearattended,
            'schooltype' : background.schooltype,
        }
        educationData.append(dic)
    educationform1 = formset_factory(EducationalBackgroundForm, extra = counts )
    if request.method == "POST":
        educationform = educationform1(request.POST,initial = educationData)
        if educationform.is_valid():
            counts = [0,0,0,0]
            totals = [0,0,0,0]
            for educations in educationform:
                totals[1] += 1
            #print(totals[1])
            if totals[1] < education.count():
                for x in range(education.count() - 1, totals[1] - 1, -1):
                    todel = education[x]
                    todel.delete()

            highest = None
            for educations in educationform:
                counts[1] += 1
                if counts[1] <= education.count():
                    cureducation = education[counts[1] - 1]
                    if educations['schoolname'].value() and educations['startingyearattended'].value() and educations['endingyearattended'].value():
                        if educationform[0] == educations:
                            highest = educations.cleaned_data['highestdegree']
                        cureducation.highestdegree= highest
                        cureducation.schoolname=educations.cleaned_data['schoolname']
                        cureducation.startingyearattended=educations.cleaned_data['startingyearattended']
                        cureducation.endingyearattended=educations.cleaned_data['endingyearattended']
                        cureducation.schooltype=educations.cleaned_data['schooltype']
                        cureducation.save()
                else:
                    if educations['schoolname'].value() and educations['startingyearattended'].value() and educations['endingyearattended'].value():
                        if educationform[0] == educations:
                            highest = educations.cleaned_data['highestdegree']
                        EducationalBackground.objects.create(
                        informationid=information,
                        highestdegree= highest,
                        schoolname=educations.cleaned_data['schoolname'],
                        startingyearattended=educations.cleaned_data['startingyearattended'],
                        endingyearattended=educations.cleaned_data['endingyearattended'],
                        schooltype=educations.cleaned_data['schooltype']
                        )
            return redirect('/home/employee/' + str(empid))
    else:

        educationform = educationform1(initial = educationData)
    context = {
    'form': educationform,}
    return render(request, 'loginapp/edit/5.html',context)

@login_required
def editfamily(request,empid):
    employee = get_object_or_404(Employee, employeeid=empid)
    information = employee.informationid   
    family = FamilyMemberBackground.objects.filter(informationid=information)
    familyData = []
    counts = 1
    for member in family:
        counts = 0
        dic = {
            'membername': member.membername,
            'memberage'  : member.memberage,
            'memberrelationship'   : member.memberrelationship,
            'memberoccupation' : member.memberoccupation,
        }
        familyData.append(dic)
    familyform1 = formset_factory(FamilyMemberBackgroundForm, extra = counts )
    if request.method == "POST":
        familyform = familyform1(request.POST,initial = familyData)
        if familyform.is_valid():
            counts = [0,0,0,0]
            totals = [0,0,0,0]
            for fammembers in familyform:
                totals[2] += 1

            if totals[2] < family.count():
                for x in range(family.count() - 1, totals[2] - 1, -1):
                    todel = family[x]
                    todel.delete()

            for fammembers in familyform:
                counts[2] += 1
                if counts[2] <= family.count():
                    if fammembers['membername'].value():
                        curfam = family[counts[2] - 1]
                        curfam.membername= fammembers.cleaned_data['membername']
                        curfam.memberage= fammembers.cleaned_data['memberage']
                        curfam.memberrelationship= fammembers.cleaned_data['memberrelationship']
                        curfam.memberoccupation= fammembers.cleaned_data['memberoccupation']
                        curfam.save()

                else:
                    if fammembers['membername'].value():
                        FamilyMemberBackground.objects.create(
                            informationid= information,
                            membername= fammembers.cleaned_data['membername'],
                            memberage= fammembers.cleaned_data['memberage'],
                            memberrelationship= fammembers.cleaned_data['memberrelationship'],
                            memberoccupation= fammembers.cleaned_data['memberoccupation']
                        )
            return redirect('/home/employee/' + str(empid))
    else:
        
        familyform = familyform1(initial = familyData)
    context = {
    'form': familyform,}
    return render(request, 'loginapp/edit/6.html',context)

@login_required
def editspouse(request,empid):
    employee = get_object_or_404(Employee, employeeid=empid)
    information = employee.informationid
    emergency = EmergencyDetails.objects.get(pk = information.emergencycontactnumber.emergencycontactnumber)
    spouse = information.spouseid
    spouseData = {}
    if spouse:
        spouseData['spousename'] = spouse.spousename
        spouseData['spousecompany'] = spouse.spousecompany
        spouseData['spousecompanyaddress'] = spouse.spousecompanyaddress
        spouseData['numberofchildren'] = spouse.numberofchildren
    spouseform = SpouseForm(spouseData)
    if request.method == "POST":
        spouseform = SpouseForm(request.POST)
        if spouseform.is_valid():
            if spouse:
                spouse.spousename = spouseform.cleaned_data['spousename']
                spouse.spousecompany = spouseform.cleaned_data['spousecompany']
                spouse.spousecompanyaddress = spouseform.cleaned_data['spousecompanyaddress']
                spouse.numberofchildren = spouseform.cleaned_data['numberofchildren']
                spouse.save()
            elif spouseform.cleaned_data['spousename'] != None and spouseform.cleaned_data['spousecompany'] != None and spouseform.cleaned_data['numberofchildren'] != None:
                temp = SpouseBackground.objects.create(spousename = spouseform.cleaned_data['spousename'], spousecompany = spouseform.cleaned_data['spousecompany'], spousecompanyaddress = spouseform.cleaned_data['spousecompanyaddress'], numberofchildren = spouseform.cleaned_data['numberofchildren'])
                information.spouseid = temp
            return redirect('/home/employee/' + str(empid))
    context = {
    'form': spouseform,}
    return render(request, 'loginapp/edit/7.html',context)


@login_required
def editchild(request,empid):
    employee = get_object_or_404(Employee, employeeid=empid)
    information = employee.informationid
    children = ChildBackground.objects.filter(informationid=information)
    childrenData = []
    counts = 1
    for child in children:
        counts = 0
        dic = {
            'childname': child.childname,
            'childage'  : child.childage,
            'childoccupation'   : child.childoccupation,
        }
        childrenData.append(dic)
    childform1 = formset_factory(ChildBackgroundForm, extra = counts)
    if request.method == "POST":
        childform = childform1(request.POST,initial = childrenData)
        if childform.is_valid():
            counts = [0,0,0,0]
            totals = [0,0,0,0]
            for x in childform:
                totals[3] += 1

            if totals[3] < children.count():
                for x in range(children.count() - 1, totals[3] - 1, -1):
                    todel = children[x]
                    todel.delete()

            for famchild in childform:
                counts[3] += 1
                if counts[3] <= children.count():
                    if famchild['childname'].value():
                        curchild = children[counts[3] - 1]
                        curchild.childname= famchild.cleaned_data['childname']
                        curchild.childage= famchild.cleaned_data['childage']
                        curchild.childoccupation= famchild.cleaned_data['childoccupation']
                        curchild.save()
                else:
                    if famchild['childname'].value():
                        ChildBackground.objects.create(
                        childname= famchild.cleaned_data['childname'],
                        childage= famchild.cleaned_data['childage'],
                        childoccupation= famchild.cleaned_data['childoccupation'],
                        informationid= information
                        )
            return redirect('/home/employee/' + str(empid))
    else:
        childform = childform1(initial = childrenData)
    context = {
    'form': childform,}
    return render(request, 'loginapp/edit/8.html',context)





@login_required
def viewtest(request):
    employees = Employee.objects.all().order_by('employeeid')
    context = {
    'employees': employees,
    }
    return render(request, 'loginapp/viewtest.html',context)

@login_required
def viewtest_awards(request):
    documents = []#Document.objects.filter(memoreferencenumber__recordtype='Award')
    for award in Awards.objects.all().iterator():
        rec = award.amemoreferencenumber
        document = Document.objects.get(memoreferencenumber = rec)
        if document.documenthide == 0:
            customdatadic = {
                'documentid' : document.documentid,
                'recordname' : rec.recordname,
                'issuingbranch' : rec.issuingbranch,
                'memoreferencenumber' : rec.memoreferencenumber,
                'issuingdepartment' : rec.issuingdepartment,
                'employeeid' : document.employeeid.employeeid,
                'employeename' : document.employeeid.informationid.employeename,
                'documentlink' : document.documentlink,
                'author' : document.author,
                'awardtype' : award.awardtype,
            }
            documents.append(customdatadic)
    if request.method == "POST":
        sortbytype = request.POST.get('sorted')
        searchempid = request.POST.get('searchempid')
        searchempname = request.POST.get('searchempname')
        searchrecordname = request.POST.get('searchrecordname')
        searchmemoreferencenumber = request.POST.get('searchmemoreferencenumber')
        searchissuingbranch = request.POST.get('searchissuingbranch')
        searchissuingdepartment = request.POST.get('searchissuingdepartment')
        searchawardtype = request.POST.get('searchawardtype')
        if searchempid != '' or searchempid != None:
            if searchempid.isdecimal():
                documents = [d for d in documents if int(searchempid) == d['employeeid']]


        if searchempname != '' :
            if searchempname.replace(' ', '').replace('.',' ').isalpha(): ## This other condition can be removed, as long as it isn't empty is what it matters
                documents = [d for d in documents if searchempname.lower() in d['employeename'].lower()]


        if searchrecordname != '' :
            if searchrecordname.replace(' ', '').replace('.',' ').isalpha():
                documents = [d for d in documents if searchrecordname.lower() in d['recordname'].lower()]


        if searchmemoreferencenumber != '' or searchmemoreferencenumber != None:
            if searchmemoreferencenumber.isdecimal():
                documents = [d for d in documents if int(searchmemoreferencenumber) == d['memoreferencenumber']]



        if searchissuingbranch != '':
            if searchissuingbranch.replace(' ', '').replace('.',' ').isalpha():
                documents = [d for d in documents if searchissuingbranch.lower() in d['issuingbranch'].lower()]
        
        if searchissuingdepartment != '':
            if searchissuingdepartment.replace(' ', '').replace('.',' ').isalpha():
                documents = [d for d in documents if searchissuingdepartment.lower() in d['issuingdepartment'].lower()]

        if searchawardtype != '':
            if searchawardtype.replace(' ', '').replace('.',' ').isalpha():
                documents = [d for d in documents if searchawardtype.lower() in d['awardtype'].lower()]

        documents = sorted(documents, key = itemgetter(sortbytype))

        context = {
        'documents': documents,
        }

        return render(request, 'loginapp/viewtest_awards.html',context)
    
    documents = sorted(documents, key = itemgetter('employeename'))
    context = {
    'documents': documents,
    }
    return render(request, 'loginapp/viewtest_awards.html',context)

@login_required
def awardsprofile(request, did):
    document =Document.objects.get(documentid = did)
    rec = document.memoreferencenumber
    
    award = get_object_or_404(Awards, amemoreferencenumber = rec)

    
    context = {
    'document': document,
    'record': rec,
    'award': award,
    }
    return render(request, 'loginapp/awardprof.html',context)

@login_required
def disciplineprofile(request, did):
    document =Document.objects.get(documentid = did)
    rec = document.memoreferencenumber
    
    noac = get_object_or_404(Noac, nmemoreferencenumber = rec)
    storageholder = None
    if noac.noactype == 'Timekeeping':
            storageholder =  Timekeeping.objects.get(tnmemoreferencenumber = noac)
            
    elif noac.noactype== 'AWOL':
            storageholder = Awol.objects.get(wnmemoreferencenumber = noac)
            
    elif noac.noactype== 'Major':
            storageholder = MajorOffense.objects.get(mnmemoreferencenumber = noac )
    
    context = {
    'document': document,
    'record': rec,
    'noac': noac,
    'type': storageholder
    }
    return render(request, 'loginapp/disciplineprof.html',context)



def editawardrecord(request,did):
    document = get_object_or_404(Document, documentid =did)
    rec = document.memoreferencenumber
    award = get_object_or_404(Awards, amemoreferencenumber = rec)

    requiredData = {
    'memoreferencenumber' :rec.memoreferencenumber,
    'recordname': rec.recordname,
    'memodate': rec.memodate,
    'recorddescription' : rec.recorddescription,
    'awardissuer' : award.awardissuer,
    'issuingbranch' : rec.issuingbranch,
    'issuingdepartment': rec.issuingdepartment,
    'awardpurpose': award.awardpurpose,
    'awardtype' : award.awardtype,
    }
    recordform = AwardsRecordEdit(initial = requiredData)
    if request.method == "POST":
        recordform = AwardsRecordEdit(request.POST, initial = requiredData)
        if recordform.is_valid():
            award.awardtype = recordform.cleaned_data['awardtype']
            award.awardpurpose = recordform.cleaned_data['awardpurpose']
            award.awardissuer = recordform.cleaned_data['awardissuer']
            rec.issuingbranch =recordform.cleaned_data['issuingbranch']
            rec.issuingdepartment = recordform.cleaned_data['issuingdepartment']
            rec.recorddescription = recordform.cleaned_data['recorddescription']
            rec.memodate = recordform.cleaned_data['memodate']
            rec.recordname = recordform.cleaned_data['recordname']
            rec.memoreferencenumber = recordform.cleaned_data['memoreferencenumber']
            
            rec.save()
            award.amemoreferencenumber = rec
            award.save()
            document.memoreferencenumber = rec
            document.save()
            return redirect('/home/awards/' + str(did))
    context = {
    'form': recordform,}
    return render(request, 'loginapp/awardrecordedit.html',context)





@login_required
def viewtest_discipline(request):
    documents = []#Document.objects.filter(memoreferencenumber__recordtype='Discipline')
    for doc in Noac.objects.all().iterator():
        rec = doc.nmemoreferencenumber
        document = Document.objects.get(memoreferencenumber = rec)
        if document.documenthide == 0:
            customdatadic = {
                'documentid' : document.documentid,
                'recordname' : rec.recordname,
                'issuingbranch' : rec.issuingbranch,
                'memoreferencenumber' : rec.memoreferencenumber,
                'issuingdepartment' : rec.issuingdepartment,
                'employeeid' : document.employeeid.employeeid,
                'employeename' : document.employeeid.informationid.employeename,
                'documentlink' : document.documentlink,
                'author' : document.author,
                'noactype' : doc.noactype,
            }
            documents.append(customdatadic)
    
    
    if request.method == "POST":
        sortbytype = request.POST.get('sorted')
        searchempid = request.POST.get('searchempid')
        searchempname = request.POST.get('searchempname')
        searchrecordname = request.POST.get('searchrecordname')
        searchmemoreferencenumber = request.POST.get('searchmemoreferencenumber')
        searchissuingbranch = request.POST.get('searchissuingbranch')
        searchissuingdepartment = request.POST.get('searchissuingdepartment')
        searchNOACtype = request.POST.get('searchNOACtype')
        if searchempid != '' or searchempid != None:
            if searchempid.isdecimal():
                documents = [d for d in documents if int(searchempid) == d['employeeid']]


        if searchempname != '' :
            if searchempname.replace(' ', '').replace('.',' ').isalpha(): ## This other condition can be removed, as long as it isn't empty is what it matters
                documents = [d for d in documents if searchempname.lower() in d['employeename'].lower()]


        if searchrecordname != '' :
            if searchrecordname.replace(' ', '').replace('.',' ').isalpha():
                documents = [d for d in documents if searchrecordname.lower() in d['recordname'].lower()]


        if searchmemoreferencenumber != '' or searchmemoreferencenumber != None:
            if searchmemoreferencenumber.isdecimal():
                documents = [d for d in documents if int(searchmemoreferencenumber) == d['memoreferencenumber']]



        if searchissuingbranch != '':
            if searchissuingbranch.replace(' ', '').replace('.',' ').isalpha():
                documents = [d for d in documents if searchissuingbranch.lower() in d['issuingbranch'].lower()]
        
        if searchissuingdepartment != '':
            if searchissuingdepartment.replace(' ', '').replace('.',' ').isalpha():
                documents = [d for d in documents if searchissuingdepartment.lower() in d['issuingdepartment'].lower()]

        if searchNOACtype != '':
            documents = [d for d in documents if searchNOACtype.lower() in d['noactype'].lower()]

        documents = sorted(documents, key = itemgetter(sortbytype))

        context = {
        'documents': documents,
        }

        return render(request, 'loginapp/viewtest_discipline.html',context)
    
    documents = sorted(documents, key = itemgetter('employeename'))
    context = {
    'documents': documents,
    }
    return render(request, 'loginapp/viewtest_discipline.html',context)

@login_required
def home(request):
    return render(request, 'loginapp/home.html')



@login_required
def viewreport(request):
    documents = Document.objects.filter(employeeid__deletehide = 0).filter(documenthide = 0)
    employeedocu = Document.objects.filter(Q(memoreferencenumber = None) , Q(employeeid__deletehide=0), Q(documenthide=0))
    awards = Document.objects.filter(Q(memoreferencenumber__recordtype='Award'),Q(employeeid__deletehide=0), Q(documenthide=0))
    discipline = Document.objects.filter(Q(memoreferencenumber__recordtype='Discipline'), Q(employeeid__deletehide=0), Q(documenthide=0))
    reportlink = '/home/genreport'
    context = {
    'documentcount' : documents.count(),
    'employeedocucount' : employeedocu.count(),
    'awardscount' : awards.count(),
    'disciplinecount' : discipline.count(),
    'documents': documents,
    'reportlink': reportlink
    }
    if request.method == "POST":
        sortbytype = request.POST.get('sorted')
        searchempid = request.POST.get('searchempid')
        searchempname = request.POST.get('searchempname')
        searchbranch = request.POST.get('searchbranch')
        searchregion = request.POST.get('searchregion')
        searchdept = request.POST.get('searchdept')
        searchposition = request.POST.get('searchposition')
        searchempstatus = request.POST.get('searchempstatus')
        employees = Employee.objects.filter(deletehide = 0)
        passed = {}

        if searchempid != '' or searchempid != None:
            if searchempid.isdecimal():
                employees = employees.filter(employeeid = int(searchempid))
                passed['empid'] = searchempid
            


        if searchempname != '' :
            if searchempname.replace(' ', '').replace('.',' ').isalpha(): ## This other condition can be removed, as long as it isn't empty is what it matters
                employees = employees.filter(informationid__employeename__icontains = searchempname)
                passed['empname'] = searchempname


        if searchbranch != '' :
            if searchbranch.replace(' ', '').replace('.',' ').isalpha():
                employees = employees.filter(branch__branch__icontains = searchbranch)
                passed['branch'] = searchbranch


        if searchregion != '' :
            if searchregion.replace(' ', '').replace('.',' ').isalpha():
                employees = employees.filter(branch__region__icontains = searchregion)
                passed['region'] = searchregion


        if searchdept != '':
            if searchdept.replace(' ', '').replace('.',' ').isalpha():
                employees = employees.filter(jobid__department__icontains = searchdept)
                passed['dept'] = searchdept


        if searchposition != '':
            if searchposition.replace(' ', '').replace('.',' ').isalpha():
                employees = employees.filter(jobid__position__icontains = searchposition)
                passed['position'] = searchposition

        if searchempstatus != '':
                employees = employees.filter(employmentstatus__icontains = searchempstatus)
                passed['empstatus'] = searchempstatus

        
        
        documents = documents.filter(employeeid__in= employees )


        if sortbytype == 'employeename' :
            documents = documents.order_by('employeeid__informationid__'+ sortbytype)
        elif sortbytype == 'employeeid':
            documents = documents.order_by('employeeid__'+ sortbytype)
        else:
            documents = documents.order_by(sortbytype)
        employeedocu = Document.objects.filter(Q(memoreferencenumber = None) , Q(employeeid__deletehide=0), Q(documenthide=0)).filter(employeeid__in= employees )
        awards = Document.objects.filter(Q(memoreferencenumber__recordtype='Award'),Q(employeeid__deletehide=0), Q(documenthide=0)).filter(employeeid__in= employees )
        discipline = Document.objects.filter(Q(memoreferencenumber__recordtype='Discipline'), Q(employeeid__deletehide=0), Q(documenthide=0)).filter(employeeid__in= employees )
        request.session['filterdata'] = passed
        context = {
        'documentcount' : documents.count(),
        'employeedocucount' : employeedocu.count(),
        'awardscount' : awards.count(),
        'disciplinecount' : discipline.count(),
        'documents': documents,
        'reportlink': reportlink
        }

        return render(request, 'loginapp/viewreport.html', context)
    request.session['filterdata'] = {}
    return render(request, 'loginapp/viewreport.html', context)







@login_required
def download(request,did):
    document = get_object_or_404(Document, documentid=did)
    file_name = document.documentname
    path_to_file = document.documentlink
    file = open(path_to_file,'rb').read()
    response = HttpResponse(file,content_type='application/force-download')
    response['Content-Disposition'] = 'attachment; filename=%s' % smart_str(file_name)
    return response

@login_required
def document(request,did):
    document = get_object_or_404(Document, documentid=did)
    path_to_file = document.documentlink
    file = open(path_to_file,'rb').read()
    response = HttpResponse(file)
    return response


@login_required
def genreport(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'General'
    employees = Employee.objects.filter(deletehide=0)
    filterdata = request.session['filterdata']
    
    if 'empid' in filterdata:
        employees = employees.filter(employeeid = int(filterdata['empid']))


    if 'empname' in filterdata:
        employees = employees.filter(informationid__employeename__icontains = filterdata['empname'])



    if 'branch' in filterdata:
        
        employees = employees.filter(branch__branch__icontains = filterdata['branch'])
        


    if 'region' in filterdata:
        
        employees = employees.filter(branch__region__icontains = filterdata['region'])
           


    if 'dept' in filterdata:
       
        employees = employees.filter(jobid__department__icontains = filterdata['dept'])
           


    if 'position' in filterdata:
        
        employees = employees.filter(jobid__position__icontains = filterdata['position'])
           

    if 'empstatus' in filterdata:
        employees = employees.filter(employmentstatus__icontains = filterdata['empstatus'])
           
    documents = Document.objects.filter(employeeid__deletehide = 0).filter(documenthide = 0).filter(employeeid__in= employees )
    employeedocu = Document.objects.filter(Q(memoreferencenumber = None) , Q(employeeid__deletehide=0), Q(documenthide=0)).filter(employeeid__in= employees )
    awards = Document.objects.filter(Q(memoreferencenumber__recordtype='Award'),Q(employeeid__deletehide=0), Q(documenthide=0)).filter(employeeid__in= employees )
    discipline = Document.objects.filter(Q(memoreferencenumber__recordtype='Discipline'), Q(employeeid__deletehide=0), Q(documenthide=0)).filter(employeeid__in= employees )
    curlevel = 2
    ## Handle Main Sheet
    ws['A1'] = 'Total Documents:'
    ws['A2'] = 'Employee Documents:'
    ws['A3'] = 'Awards & Recognition:'
    ws['A4'] = 'Discipline Management:'
    ws['E1'] = documents.count()
    ws['E2'] = employeedocu.count()
    ws['E3'] = awards.count()
    ws['E4'] = discipline.count()
    ws['A6'] = 'CHECK OTHER CORRESPONDING SHEETS FOR REPORTS'
    ws['A6'].font = Font(size=64)

    EmployeeSheet = wb.create_sheet('Employees')
    EmployeeRecSheet = wb.create_sheet('Employee Records')
    AwardsSheet = wb.create_sheet('Awards Records')
    DisciplineSheet = wb.create_sheet('Discipline Records')

    #Handle Employee Sheet

    EmployeeSheet['A1'] = 'Link'
    EmployeeSheet['B1'] ='Employee ID'
    EmployeeSheet['C1'] ='Employee Name'
    EmployeeSheet['D1'] ='Department'
    EmployeeSheet['E1'] ='Position'
    EmployeeSheet['F1'] ='Employment Status'
    EmployeeSheet['G1'] ='Start Date'
    EmployeeSheet['H1'] ='End Date'
    EmployeeSheet['I1'] ='Salary (PHP)'
    EmployeeSheet['J1'] ='Salary Type'

    for employee in employees:
        EmployeeSheet['A' + str(curlevel)] = '127.0.0.1/home/employee/'	+ str(employee.employeeid) + '/'
        EmployeeSheet['B'+ str(curlevel)] =employee.employeeid
        EmployeeSheet['C'+ str(curlevel)] =employee.informationid.employeename
        EmployeeSheet['D'+ str(curlevel)] = employee.jobid.department
        EmployeeSheet['E'+ str(curlevel)] = employee.jobid.position
        EmployeeSheet['F'+ str(curlevel)] = employee.employmentstatus
        EmployeeSheet['G'+ str(curlevel)] = str(employee.startdate)
        EmployeeSheet['H'+ str(curlevel)] =	str(employee.enddate)
        EmployeeSheet['I'+ str(curlevel)] = employee.salary
        EmployeeSheet['J'+ str(curlevel)] = employee.salarytype
        curlevel += 1

    curlevel = 2

    #Handle Employee Documents

    EmployeeRecSheet['A1'] = 'Link'
    EmployeeRecSheet['B1'] ='Employee ID'
    EmployeeRecSheet['C1'] ='Document ID'
    EmployeeRecSheet['D1'] ='Document Name'
    EmployeeRecSheet['E1'] ='Date Created'
    EmployeeRecSheet['F1'] ='Author'

    for docu in employeedocu:
        EmployeeRecSheet['A' + str(curlevel)] = docu.documentlink
        EmployeeRecSheet['B'+ str(curlevel)] =docu.employeeid.employeeid
        EmployeeRecSheet['C'+ str(curlevel)] =docu.documentid
        EmployeeRecSheet['D'+ str(curlevel)] = docu.documentname
        EmployeeRecSheet['E'+ str(curlevel)] = str(docu.dateandtimecreated)
        EmployeeRecSheet['F'+ str(curlevel)] = docu.author
        curlevel += 1

    curlevel = 2

    #Handle Employee Documents

    AwardsSheet['A1'] = 'Link'
    AwardsSheet['B1'] ='Employee ID'
    AwardsSheet['C1'] ='Memo Reference Number'
    AwardsSheet['D1'] ='Document Name'
    AwardsSheet['E1'] ='Date Created'
    AwardsSheet['F1'] ='Author'

    for docu in awards:
        AwardsSheet['A' + str(curlevel)] = docu.documentlink
        AwardsSheet['B'+ str(curlevel)] =docu.employeeid.employeeid
        AwardsSheet['C'+ str(curlevel)] =docu.memoreferencenumber.memoreferencenumber
        AwardsSheet['D'+ str(curlevel)] = docu.documentname
        AwardsSheet['E'+ str(curlevel)] = str(docu.dateandtimecreated)
        AwardsSheet['F'+ str(curlevel)] = docu.author
        curlevel += 1

    curlevel = 2

    #Handle Employee Documents

    DisciplineSheet['A1'] = 'Link'
    DisciplineSheet['B1'] ='Employee ID'
    DisciplineSheet['C1'] ='Memo Reference Number'
    DisciplineSheet['D1'] ='Document Name'
    DisciplineSheet['E1'] ='Date Created'
    DisciplineSheet['F1'] ='Author'

    for docu in discipline:
        DisciplineSheet['A' + str(curlevel)] = docu.documentlink
        DisciplineSheet['B'+ str(curlevel)] =docu.employeeid.employeeid
        DisciplineSheet['C'+ str(curlevel)] =docu.memoreferencenumber.memoreferencenumber
        DisciplineSheet['D'+ str(curlevel)] = docu.documentname
        DisciplineSheet['E'+ str(curlevel)] = str(docu.dateandtimecreated)
        DisciplineSheet['F'+ str(curlevel)] = docu.author
        curlevel += 1
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')

    return response




def uploademployeerecord(request,empid):
    employee = get_object_or_404(Employee, employeeid=empid)


    if request.method == "POST":

        recordform = EmployeeDocument(request.POST, request.FILES)
        if recordform.is_valid():
            time = datetime.datetime.now()
            user = request.user.username
            recordfile = request.FILES['recordfile']
            dest = 'media/employee/' + str(empid) + '/employeerecords'
            fs = FileSystemStorage(location = dest, base_url = dest)
            filename = fs.save(recordfile.name,recordfile)
            fileurl = fs.url(filename)
            Document.objects.create(
            documentname = recordfile.name  ,
            dateandtimecreated = time ,
            documentlink = fileurl,
            author = user,
            dateandtimelastedited = time  ,
            recenteditor = user  ,
            employeeid =employee  ,
            preparedby = recordform.cleaned_data['preparedby']  ,
            preparationdate = recordform.cleaned_data['preparationdate']  ,
            notedby = recordform.cleaned_data['notedby'] ,
            noteddate = recordform.cleaned_data['noteddate']  ,
            approvedby = recordform.cleaned_data['approvedby'] ,
            approveddate = recordform.cleaned_data['approveddate']  ,
            receivedby =recordform.cleaned_data['receivedby']  ,
            receiveddate = recordform.cleaned_data['receiveddate'],
            documenthide = 0
            )
            return redirect('/home/employee/' + str(empid))
    
    recordform = EmployeeDocument()
    return render(request, 'loginapp/recordtest.html',{'record':recordform})


def editedocument(request,empid,did):
    doc = get_object_or_404(Document, documentid=did)
    employee = get_object_or_404(Employee, employeeid=empid)
    initialdata = {
    'preparedby' : doc.preparedby,
    'preparationdate' : doc.preparationdate,
    'notedby' : doc.notedby,
    'noteddate' : doc.noteddate,
    'approvedby' : doc.approvedby,
    'approveddate' : doc.approveddate,
    'receivedby': doc.receivedby,
    'receiveddate' : doc.receiveddate,
    }

    if request.method == "POST":

        recordform = EditDocument(request.POST, request.FILES, initial = initialdata)
        if recordform.is_valid():
            time = datetime.datetime.now()
            user = request.user.username
            if request.FILES['recordfile']:
                oldfilepath = os.path.splitext(doc.documentlink)
                newfilename = oldfilepath[0] + datetime.datetime.now().strftime("%Y%m%d-%H%M%S") + oldfilepath[1]
                os.rename(doc.documentlink,newfilename)
                recordfile = request.FILES['recordfile']
                dest = 'media/employee/' + str(empid) + '/employeerecords'
                fs = FileSystemStorage(location = dest, base_url = dest)
                filename = fs.save(recordfile.name,recordfile)
                fileurl = fs.url(filename)
                doc.documentname = recordfile.name
                doc.documentlink = fileurl

            doc.dateandtimelastedited = time  
            doc.recenteditor = user  
            doc.preparedby = recordform.cleaned_data['preparedby']  
            doc.preparationdate = recordform.cleaned_data['preparationdate']  
            doc.notedby = recordform.cleaned_data['notedby'] 
            doc.noteddate = recordform.cleaned_data['noteddate']  
            doc.approvedby = recordform.cleaned_data['approvedby'] 
            doc.approveddate = recordform.cleaned_data['approveddate']  
            doc.receivedby =recordform.cleaned_data['receivedby']  
            doc.receiveddate = recordform.cleaned_data['receiveddate']
            doc.save()
            return redirect('/home/employee/' + str(empid))
    
    recordform = EditDocument(initial = initialdata)
    return render(request, 'loginapp/recordtest.html',{'record':recordform})

def editadocument(request,did):
    doc = get_object_or_404(Document, documentid=did)
    employee = doc.employeeid
    empid = employee.employeeid
    initialdata = {
    'preparedby' : doc.preparedby,
    'preparationdate' : doc.preparationdate,
    'notedby' : doc.notedby,
    'noteddate' : doc.noteddate,
    'approvedby' : doc.approvedby,
    'approveddate' : doc.approveddate,
    'receivedby': doc.receivedby,
    'receiveddate' : doc.receiveddate,
    }

    if request.method == "POST":

        recordform = EditDocument(request.POST, request.FILES, initial = initialdata)
        if recordform.is_valid():
            time = datetime.datetime.now()
            user = request.user.username
            if request.FILES['recordfile']:
                oldfilepath = os.path.splitext(doc.documentlink)
                newfilename = oldfilepath[0] + datetime.datetime.now().strftime("%Y%m%d-%H%M%S") + oldfilepath[1]
                os.rename(doc.documentlink,newfilename)
                recordfile = request.FILES['recordfile']
                dest = 'media/employee/' + str(empid) + '/awards'
                fs = FileSystemStorage(location = dest, base_url = dest)
                filename = fs.save(recordfile.name,recordfile)
                fileurl = fs.url(filename)
                doc.documentname = recordfile.name
                doc.documentlink = fileurl

            doc.dateandtimelastedited = time  
            doc.recenteditor = user  
            doc.preparedby = recordform.cleaned_data['preparedby']  
            doc.preparationdate = recordform.cleaned_data['preparationdate']  
            doc.notedby = recordform.cleaned_data['notedby'] 
            doc.noteddate = recordform.cleaned_data['noteddate']  
            doc.approvedby = recordform.cleaned_data['approvedby'] 
            doc.approveddate = recordform.cleaned_data['approveddate']  
            doc.receivedby =recordform.cleaned_data['receivedby']  
            doc.receiveddate = recordform.cleaned_data['receiveddate']
            doc.save()
            return redirect('/home/awards/' + str(did))
    
    recordform = EditDocument(initial = initialdata)
    return render(request, 'loginapp/recordtest.html',{'record':recordform})



AWARDSFORMS = [('1',  AwardsRecord ),
('2',  RecordDocument ),
]
AWARDSTEMPLATES = {
'1' : 'loginapp/awards/record.html',
'2' : 'loginapp/awards/document.html',

}
class AwardsWizard(SessionWizardView):

    file_storage = FileSystemStorage(location='/media/temporary')
    def get_template_names(self):
        return [AWARDSTEMPLATES[self.steps.current]]

    def done(self, form_list, **kwargs):
        #data = [form.cleaned_data for form in form_list] ##Ignore form list since sorting it out with ifs is a pain
        record = self.get_cleaned_data_for_step('1')
        document = self.get_cleaned_data_for_step('2')
        destinationemployee = record['recordfor']
        time = datetime.datetime.now()
        user = self.request.user.username
        recordfile = self.get_cleaned_data_for_step('2')['recordfile']
        dest = 'media/employee/' + str(destinationemployee.employeeid) + '/awards'
        fs = FileSystemStorage(location = dest, base_url = dest)
        filename = fs.save(recordfile.name,recordfile)
        fileurl = fs.url(filename)

        recstorage = Record.objects.create(
            memoreferencenumber = record['memoreferencenumber'],
            recordname = record['recordname'],
            memodate = record['memodate'],
            issuingbranch = record['issuingbranch'],
            recorddescription = record['recorddescription'],
            recordtype = 'Award',
            issuingdepartment = record['issuingdepartment'],


        )

        award = Awards.objects.create(
            amemoreferencenumber = recstorage,
            awardissuer = record['awardissuer'],
            awardpurpose = record['awardpurpose'],
            awardtype = record['awardtype']

        )

        Document.objects.create(
        documentname = recordfile.name  ,
        dateandtimecreated = time ,
        documentlink = fileurl,
        author = user,
        dateandtimelastedited = time  ,
        recenteditor = user  ,
        employeeid =destinationemployee  ,
        preparedby = document['preparedby']  ,
        preparationdate = document['preparationdate']  ,
        notedby = document['notedby'] ,
        noteddate = document['noteddate']  ,
        approvedby = document['approvedby'] ,
        approveddate = document['approveddate']  ,
        receivedby =document['receivedby']  ,
        receiveddate = document['receiveddate'],
        documenthide = 0,
        memoreferencenumber = recstorage)















        return redirect('viewtest_awards')








    def render_goto_step(self, goto_step, **kwargs):


        form1 = self.get_form(self.storage.current_step, data=self.request.POST,files=self.request.FILES)
        if form1.is_valid:
            self.storage.set_step_data(self.storage.current_step, self.process_step(form1))
            self.storage.set_step_files(self.storage.current_step, self.process_step_files(form1))


        ######### this is from render_goto_step method
        self.storage.current_step = goto_step

        form = self.get_form(
            data=self.storage.get_step_data(self.steps.current),
            files=self.storage.get_step_files(self.steps.current),
            )

        return self.render(form)



award_view = AwardsWizard.as_view(AWARDSFORMS)

@login_required
def uploadaward(request):
     return award_view(request)


def editdisciplinerecord(request,did):
    document = get_object_or_404(Document, documentid =did)
    rec = document.memoreferencenumber
    noac = get_object_or_404(Noac, nmemoreferencenumber = rec)
    storageholder = None

    requiredData = {
    'memoreferencenumber' :rec.memoreferencenumber,
    'recordname': rec.recordname,
    'memodate': rec.memodate,
    'recorddescription' : rec.recorddescription,
    'issuingbranch' : rec.issuingbranch,
    'issuingdepartment': rec.issuingdepartment,
    'remarks': noac.remarks,
    'noactype' : noac.noactype,
    'sanction': noac.sanction,
    }

    if noac.noactype == 'Timekeeping':
            storageholder =  Timekeeping.objects.get(tnmemoreferencenumber = noac)
            requiredData['noofoffense'] = storageholder.noofoffense
    elif noac.noactype== 'AWOL':
            storageholder = Awol.objects.get(wnmemoreferencenumber = noac)
            requiredData['noofoffense'] = storageholder.noofoffense
            requiredData['hearingdate'] = storageholder.hearingdate
    elif noac.noactype== 'Major':
            storageholder = MajorOffense.objects.get(mnmemoreferencenumber = noac )
            requiredData['hearingdate'] = storageholder.hearingdate

    

        
    
    recordform = DisciplineRecordEdit(initial = requiredData)
    if request.method == "POST":
        recordform = DisciplineRecordEdit(request.POST, initial = requiredData)
        if recordform.is_valid():
            rec.issuingbranch =recordform.cleaned_data['issuingbranch']
            rec.issuingdepartment = recordform.cleaned_data['issuingdepartment']
            rec.recorddescription = recordform.cleaned_data['recorddescription']
            rec.memodate = recordform.cleaned_data['memodate']
            rec.recordname = recordform.cleaned_data['recordname']
            rec.memoreferencenumber = recordform.cleaned_data['memoreferencenumber']
            rec.save()
            noac.remarks = recordform.cleaned_data['remarks']
            noac.noactype = recordform.cleaned_data['noactype']
            noac.sanction = recordform.cleaned_data['sanction']
            noac.nmemoreferencenumber = rec
            noac.save()
            if recordform.cleaned_data['noactype'] != noac.noactype:
                storageholder.delete()
                if recordform.cleaned_data['noactype'] == 'Timekeeping':
                    Timekeeping.objects.create(noofoffense = recordform.cleaned_data['noofoffense'],tnmemoreferencenumber = noac)
                elif recordform.cleaned_data['noactype'] == 'AWOL':
                    Awol.objects.create(
                    hearingdate = recordform.cleaned_data['hearingdate'],
                    noofoffense = recordform.cleaned_data['noofoffense'],
                    wnmemoreferencenumber = noac)
                elif recordform.cleaned_data['noactype'] == 'Major':
                    MajorOffense.objects.create(hearingdate = recordform.cleaned_data['hearingdate'],mnmemoreferencenumber = noac )
            else:
                if recordform.cleaned_data['noactype'] == 'Timekeeping':
                    storageholder.noofoffense =  recordform.cleaned_data['noofoffense']
                    storageholder.tnmemoreferencenumber =  noac
                    storageholder.save()

                elif recordform.cleaned_data['noactype'] == 'AWOL':
                    storageholder.noofoffense =  recordform.cleaned_data['noofoffense']
                    storageholder.hearingdate =  recordform.cleaned_data['hearingdate']
                    storageholder.wnmemoreferencenumber =  noac
                    storageholder.save()

                elif recordform.cleaned_data['noactype'] == 'Major':
                    storageholder.hearingdate =  recordform.cleaned_data['hearingdate']
                    storageholder.mnmemoreferencenumber =  noac
                    storageholder.save()


            document.memoreferencenumber = rec
            document.save()
            return redirect('/home/discipline/' + str(did))
    context = {
    'form': recordform,}
    return render(request, 'loginapp/disciplinerecordedit.html',context)




def editddocument(request,did):
    doc = get_object_or_404(Document, documentid=did)
    employee = doc.employeeid
    empid = employee.employeeid
    initialdata = {
    'preparedby' : doc.preparedby,
    'preparationdate' : doc.preparationdate,
    'notedby' : doc.notedby,
    'noteddate' : doc.noteddate,
    'approvedby' : doc.approvedby,
    'approveddate' : doc.approveddate,
    'receivedby': doc.receivedby,
    'receiveddate' : doc.receiveddate,
    }

    if request.method == "POST":

        recordform = EditDocument(request.POST, request.FILES, initial = initialdata)
        if recordform.is_valid():
            time = datetime.datetime.now()
            user = request.user.username
            if request.FILES['recordfile']:
                oldfilepath = os.path.splitext(doc.documentlink)
                newfilename = oldfilepath[0] + datetime.datetime.now().strftime("%Y%m%d-%H%M%S") + oldfilepath[1]
                os.rename(doc.documentlink,newfilename)
                recordfile = request.FILES['recordfile']
                dest = 'media/employee/' + str(empid) + '/discipline'
                fs = FileSystemStorage(location = dest, base_url = dest)
                filename = fs.save(recordfile.name,recordfile)
                fileurl = fs.url(filename)
                doc.documentname = recordfile.name
                doc.documentlink = fileurl

            doc.dateandtimelastedited = time  
            doc.recenteditor = user  
            doc.preparedby = recordform.cleaned_data['preparedby']  
            doc.preparationdate = recordform.cleaned_data['preparationdate']  
            doc.notedby = recordform.cleaned_data['notedby'] 
            doc.noteddate = recordform.cleaned_data['noteddate']  
            doc.approvedby = recordform.cleaned_data['approvedby'] 
            doc.approveddate = recordform.cleaned_data['approveddate']  
            doc.receivedby =recordform.cleaned_data['receivedby']  
            doc.receiveddate = recordform.cleaned_data['receiveddate']
            doc.save()
            return redirect('/home/discipline/' + str(did))
    
    recordform = EditDocument(initial = initialdata)
    return render(request, 'loginapp/recordtest.html',{'record':recordform})



DISCIPLINEFORMS = [('1',  DisciplineRecord ),
('2',  RecordDocument )]
DISCIPLINETEMPLATES = {
'1' : 'loginapp/discipline/record.html',
'2' : 'loginapp/discipline/document.html',

}
class DisciplineWizard(SessionWizardView):

    file_storage = FileSystemStorage(location='/media/temporary')
    def get_template_names(self):
        return [DISCIPLINETEMPLATES[self.steps.current]]

    def done(self, form_list, **kwargs):
        #data = [form.cleaned_data for form in form_list] ##Ignore form list since sorting it out with ifs is a pain
        record = self.get_cleaned_data_for_step('1')
        document = self.get_cleaned_data_for_step('2')
        destinationemployee = record['recordfor']
        time = datetime.datetime.now()
        user = self.request.user.username
        recordfile = self.get_cleaned_data_for_step('2')['recordfile']
        dest = 'media/employee/' + str(destinationemployee.employeeid) + '/discipline'
        fs = FileSystemStorage(location = dest, base_url = dest)
        filename = fs.save(recordfile.name,recordfile)
        fileurl = fs.url(filename)

        recstorage = Record.objects.create(
            memoreferencenumber = record['memoreferencenumber'],
            recordname = record['recordname'],
            memodate = record['memodate'],
            issuingbranch = record['issuingbranch'],
            recorddescription = record['recorddescription'],
            recordtype = 'Discipline',
            issuingdepartment = record['issuingdepartment'],


        )
        noacstorage = Noac.objects.create(
        remarks = record['remarks'],
        noactype = record['noactype'],
        sanction = record['sanction'],
        nmemoreferencenumber = recstorage,)

        if record['noactype'] == 'Timekeeping':
            Timekeeping.objects.create(noofoffense = record['noofoffense'],tnmemoreferencenumber = noacstorage)
        elif record['noactype'] == 'AWOL':
            Awol.objects.create(
                hearingdate = record['hearingdate'],
                noofoffense = record['noofoffense'],
                wnmemoreferencenumber = noacstorage)
        elif record['noactype'] == 'Major':
            MajorOffense.objects.create(hearingdate = record['hearingdate'],mnmemoreferencenumber = noacstorage )
        
        Document.objects.create(
        documentname = recordfile.name  ,
        dateandtimecreated = time ,
        documentlink = fileurl,
        author = user,
        dateandtimelastedited = time  ,
        recenteditor = user  ,
        employeeid =destinationemployee  ,
        preparedby = document['preparedby']  ,
        preparationdate = document['preparationdate']  ,
        notedby = document['notedby'] ,
        noteddate = document['noteddate']  ,
        approvedby = document['approvedby'] ,
        approveddate = document['approveddate']  ,
        receivedby =document['receivedby']  ,
        receiveddate = document['receiveddate'],
        documenthide = 0,
        memoreferencenumber = recstorage)















        return redirect('viewtest_discipline')








    def render_goto_step(self, goto_step, **kwargs):


        form1 = self.get_form(self.storage.current_step, data=self.request.POST,files=self.request.FILES)
        if form1.is_valid:
            self.storage.set_step_data(self.storage.current_step, self.process_step(form1))
            self.storage.set_step_files(self.storage.current_step, self.process_step_files(form1))


        ######### this is from render_goto_step method
        self.storage.current_step = goto_step

        form = self.get_form(
            data=self.storage.get_step_data(self.steps.current),
            files=self.storage.get_step_files(self.steps.current),
            )

        return self.render(form)



discipline_view = DisciplineWizard.as_view(DISCIPLINEFORMS)

@login_required
def uploaddiscipline(request):
     return discipline_view(request)


















def testing(request):
    branch = formset_factory(BranchForm)
    if request.method == "POST":
        branch = branch(request.POST)

        if branch.is_valid():
            for branches in branch:
                EmployeeWorkLocation.objects.create(branch = branches.cleaned_data['branch'], region = branches.cleaned_data['region'])

    return render(request, 'loginapp/testing.html',{'branch' : branch})































RECORDFORMS = [('1',  EmployeeRecordForm ),
('2',  EmployeePersonalForm ),
('3',  EmergencyForm ),
('4',  formset_factory(EmploymentHistoryForm, extra = 1 ) ),
('5',  formset_factory(EducationalBackgroundForm, extra = 1) ),
('6',  formset_factory(FamilyMemberBackgroundForm, extra = 1) ),
('7',  SpouseForm ),
('8',  formset_factory(ChildBackgroundForm, extra = 1) ),
]
RECORDTEMPLATES = {
'1' : 'loginapp/create/1.html',
'2' : 'loginapp/create/2.html',
'3' : 'loginapp/create/3.html',
'4' : 'loginapp/create/4.html',
'5' : 'loginapp/create/5.html',
'6': 'loginapp/create/6.html',
'7' : 'loginapp/create/7.html',
'8' : 'loginapp/create/8.html',
}
class EmployeeWizard(SessionWizardView):

    def get_template_names(self):
        return [RECORDTEMPLATES[self.steps.current]]

    def done(self, form_list, **kwargs):
        #data = [form.cleaned_data for form in form_list] ##Ignore form list since sorting it out with ifs is a pain
        record = self.get_cleaned_data_for_step('1')
        personal = self.get_cleaned_data_for_step('2')
        emergency = self.get_cleaned_data_for_step('3')
        emphistory = self.get_cleaned_data_for_step('4')
        education = self.get_cleaned_data_for_step('5')
        family = self.get_cleaned_data_for_step('6')
        spouse = self.get_cleaned_data_for_step('7')
        child = self.get_cleaned_data_for_step('8')
        spousetemp = None
        if 'spousename' in spouse and 'spousecompany' in spouse and 'numberofchildren' in spouse:
            if spouse['spousename'] != None and spouse['spousecompany'] != None and spouse['numberofchildren'] != None:
                spousenamearg = Q(spousename__contains = spouse['spousename'])
                spousecomparg = Q(spousecompany__contains = spouse['spousecompany'])
                spousechildren = Q(numberofchildren__contains = spouse['numberofchildren'])
                spousecheck = checkmodelq(SpouseBackground, spousenamearg & spousecomparg & spousechildren)
                if spousecheck == None:
                    spousestore = SpouseBackground.objects.create(spousename = spouse['spousename'], spousecompany = spouse['spousecompany'], spousecompanyaddress = spouse['spousecompanyaddress'], numberofchildren = spouse['numberofchildren'])
                    spousetemp = spousestore
                else:
                    spousetemp = spousecheck


        emergencycheck = checkmodel(EmergencyDetails,emergencycontactnumber = emergency['emergencycontactnumber'])

        if  emergencycheck == None:
            emergencystore = EmergencyDetails.objects.create(emergencycontactnumber = emergency['emergencycontactnumber'], emergencycontactname= emergency['emergencycontactname'], emergencyrelationship= emergency['emergencyrelationship'],emergencyaddress=emergency['emergencyaddress'])
            emergency = emergencystore
        else:
            emergency = emergencycheck

        pinfostore = EmployeePersonalInfo.objects.create(emergencycontactnumber= emergency,
        employeename= personal['employeename'],
        gender = personal['gender'],
        birthdate = personal['birthdate'],
        civilstatus = personal['civilstatus'],
        citizenship = personal['citizenship'],
        religion = personal['religion'],
        bloodtype = personal['bloodtype'],
        numberofdependent = personal['numberofdependent'],
        presentaddress = personal['presentaddress'],
        permanentaddress = personal['permanentaddress'],
        contactnumber = personal['contactnumber'],
        spouseid = spousetemp)
        informationid = pinfostore

        recordstore = Employee.objects.create(
            informationid= informationid,
            employeeid=  record['employeeid'],
            branch = record['branch'],
            startdate = record['startdate'],
            enddate = record['enddate'],
            employmentstatus = record['employmentstatus'],
            salarytype = record['salarytype'],
            salary = record['salary'],
            deletehide = 0,
            jobid = record['jobid']
        )


        for histories in emphistory:
            if 'previouscompanyname' in histories :
                if histories['previouscompanyname'] != '' and histories['previouscompanyname'] != None:
                    EmploymentHistory.objects.create(
                    informationid=informationid,
                    previouscompanyname= histories['previouscompanyname'],
                    previousposition= histories['previousposition'],
                    reasonforleaving= histories['reasonforleaving'],
                    companycontactnumber= histories['companycontactnumber'],
                    withcoeorclearance= histories['withcoeorclearance']
                    )
        highest = None

        for educations in education:
            if education[0] == educations:
                    highest = educations['highestdegree']
            if 'schoolname' in educations and 'schoolname' in educations and 'endingyearattended' in educations:
                if educations['schoolname']!= None and educations['schoolname']!= '' and educations['startingyearattended']!= None and educations['endingyearattended'] != None:
                    EducationalBackground.objects.create(
                    informationid=informationid,
                    highestdegree= highest,
                    schoolname=educations['schoolname'],
                    startingyearattended=educations['startingyearattended'],
                    endingyearattended=educations['endingyearattended'],
                    schooltype=educations['schooltype']
                    )

        for fammembers in family:
            if 'membername' in fammembers :
                if fammembers['membername'] != None and fammembers['membername'] != '':
                    FamilyMemberBackground.objects.create(
                        informationid= informationid,
                        membername= fammembers['membername'],
                        memberage= fammembers['memberage'],
                        memberrelationship= fammembers['memberrelationship'],
                        memberoccupation= fammembers['memberoccupation']
                    )

        for children in child:
            if 'childname' in children :
                 if children['childname'] != None and  children['childname'] != '':
                    ChildBackground.objects.create(
                    childname= children['childname'],
                    childage= children['childage'],
                    childoccupation= children['childoccupation'],
                    informationid= informationid
                    )


        return redirect('employeeform')








    def render_goto_step(self, goto_step, **kwargs):


        form1 = self.get_form(self.storage.current_step, data=self.request.POST,files=self.request.FILES)
        if form1.is_valid:
            self.storage.set_step_data(self.storage.current_step, self.process_step(form1))
            self.storage.set_step_files(self.storage.current_step, self.process_step_files(form1))


        ######### this is from render_goto_step method
        self.storage.current_step = goto_step

        form = self.get_form(
            data=self.storage.get_step_data(self.steps.current),
            files=self.storage.get_step_files(self.steps.current),
            )

        return self.render(form)



wizard_view = EmployeeWizard.as_view(RECORDFORMS)

@login_required
def createrecord(request):
     return wizard_view(request)
   




def logoutuser(request):
    logout(request)
